from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

EXCEL_OUTPUT = Path("invoices.xlsx")
HEADERS = [
    "Vendor Name",
    "Client Name",
    "Invoice Number",
    "Invoice Date",
    "Due Date",
    "GST Number",
    "Total Amount",
    "Due Amount",
    "Currency",
    "Source File",
    "Processed At",
]

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_BORDER      = Border(
    bottom=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
)
_COL_WIDTHS  = [25, 25, 18, 14, 14, 35, 14, 14, 10, 35, 18]


def init_excel():
    """Create invoices.xlsx with headers and formatting if it doesn't exist."""
    if EXCEL_OUTPUT.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    wb.save(EXCEL_OUTPUT)
    print(f"[Excel] Created {EXCEL_OUTPUT}")


def append_to_excel(record: dict):
    """Append one invoice record with alternating row shading."""
    wb = load_workbook(EXCEL_OUTPUT)
    ws = wb.active

    # Find the last real data row (ignore any Grand Total rows)
    last_data_row = 1
    for row in ws.iter_rows(min_row=2):
        # A row counts as data if the first cell has a non-empty value
        # AND it isn't our summary label
        first_val = row[0].value
        if first_val not in (None, "", "Grand Total"):
            last_data_row = row[0].row

    next_row = last_data_row + 1

    # Clear any existing Grand Total rows below the data so they don't pile up
    for r in range(next_row, next_row + 5):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).value = None

    fill = PatternFill("solid", start_color="D6E4F0" if next_row % 2 == 0 else "FFFFFF")
    font = Font(name="Arial", size=10)

    values = [
        record.get("vendor_name", ""),
        record.get("client_name", ""),
        record.get("invoice_number", ""),
        record.get("invoice_date", ""),
        record.get("due_date", ""),
        record.get("gst_number", ""),
        record.get("total_amount", ""),
        record.get("due_amount", ""),
        record.get("currency", "INR"),
        record.get("source_file", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = Alignment(vertical="center")

    # Grand Total formula — placed 2 rows below the last data row
    total_col_idx  = HEADERS.index("Total Amount") + 1
    total_col_ltr  = ws.cell(row=1, column=total_col_idx).column_letter
    summary_row    = next_row + 2

    label_cell = ws.cell(row=summary_row, column=total_col_idx - 1, value="Grand Total")
    label_cell.font = Font(bold=True, name="Arial", size=10)

    formula_cell = ws.cell(
        row=summary_row,
        column=total_col_idx,
        value=f"=SUM({total_col_ltr}2:{total_col_ltr}{next_row})",
    )
    formula_cell.font = Font(bold=True, name="Arial", size=10)

    wb.save(EXCEL_OUTPUT)
    print(f"[Excel] Saved invoice → row {next_row}")